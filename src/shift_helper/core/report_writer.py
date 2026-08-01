"""Write selected outages into a copy of the approved new report template."""

from __future__ import annotations

import os
import re
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .events import JournalEvent

REPORT_SHEET = "Аварийные отключения ЛЭП"
HEADER_ROW = 3
FIRST_DATA_ROW = 4
EXPECTED_HEADERS = (
    "Диспетчерское наименование ЛЭП, ВЭУ и оборудования",
    "Дата, время отключения",
    "Причина",
    "Работа защит",
    "Дата, время включения в работу",
)


class ReportTemplateError(ValueError):
    """Raised when the supplied report template does not match the approved layout."""


def _copy_cell_style(source: Cell, target: Cell) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _verify_headers(ws: Worksheet) -> None:
    actual = tuple(str(ws.cell(HEADER_ROW, column).value or "").strip() for column in range(2, 7))
    if actual != EXPECTED_HEADERS:
        raise ReportTemplateError(
            f"Лист {REPORT_SHEET!r} не соответствует утверждённой карте полей: {actual!r}."
        )


def _update_title(ws: Worksheet, report_date: date) -> None:
    title = str(ws["B1"].value or "")
    replacement = report_date.strftime("%d.%m.%Y")
    if re.search(r"\d{2}\.\d{2}\.\d{4}", title):
        ws["B1"] = re.sub(r"\d{2}\.\d{2}\.\d{4}", replacement, title, count=1)


def _clear_existing(ws: Worksheet) -> None:
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        for column in range(2, 7):
            ws.cell(row, column).value = None


def _prepare_rows(ws: Worksheet, count: int) -> None:
    prototype_row = FIRST_DATA_ROW
    needed_last = FIRST_DATA_ROW + max(count, 1) - 1
    if needed_last > ws.max_row:
        ws.insert_rows(ws.max_row + 1, amount=needed_last - ws.max_row)
    prototype_height = ws.row_dimensions[prototype_row].height
    for row in range(FIRST_DATA_ROW, needed_last + 1):
        if row != prototype_row:
            for column in range(2, 7):
                _copy_cell_style(ws.cell(prototype_row, column), ws.cell(row, column))
        ws.row_dimensions[row].height = prototype_height


def _write_events(ws: Worksheet, events: list[JournalEvent]) -> None:
    _prepare_rows(ws, len(events))
    for index, event in enumerate(events, start=FIRST_DATA_ROW):
        values: tuple[object, ...] = (
            event.dispatch_name,
            event.started_at,
            event.reason,
            event.description,
            event.ended_at,
        )
        for column, value in enumerate(values, start=2):
            ws.cell(index, column).value = value
        ws.cell(index, 3).number_format = "dd.mm.yyyy hh:mm"
        ws.cell(index, 6).number_format = "dd.mm.yyyy hh:mm"


def build_emergency_report(
    *,
    template_path: str | Path,
    output_path: str | Path,
    report_date: date,
    events: list[JournalEvent],
) -> Path:
    """Create an atomic report copy and verify the published workbook."""

    template = Path(template_path).resolve()
    output = Path(output_path).resolve()
    if output in {template}:
        raise ReportTemplateError("Результат не должен перезаписывать шаблон.")
    output.parent.mkdir(parents=True, exist_ok=True)
    pending = output.with_name(f".{output.name}.pending.xlsx")
    pending.unlink(missing_ok=True)

    workbook = load_workbook(template, keep_links=False)
    try:
        if REPORT_SHEET not in workbook.sheetnames:
            raise ReportTemplateError(f"В шаблоне отсутствует лист {REPORT_SHEET!r}.")
        ws = workbook[REPORT_SHEET]
        _verify_headers(ws)
        _update_title(ws, report_date)
        _clear_existing(ws)
        _write_events(ws, events)
        workbook.save(pending)
    finally:
        workbook.close()

    verifier = load_workbook(pending, read_only=True, data_only=False, keep_links=False)
    try:
        ws = verifier[REPORT_SHEET]
        _verify_headers(ws)
        populated = sum(
            1
            for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + len(events))
            if ws.cell(row, 2).value not in (None, "")
        )
        if populated != len(events):
            raise ReportTemplateError(
                f"Проверка результата ожидала {len(events)} строк, найдено {populated}."
            )
    finally:
        verifier.close()

    os.replace(pending, output)
    return output

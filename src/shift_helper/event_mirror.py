"""Atomic derived .xlsx mirror for the approved ЖС journal form."""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from threading import Lock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select, text
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

from .models import Event

EVENT_MIRROR_FILENAME = "Журнал событий.xlsx"
EVENT_MIRROR_SHEET = "ЖС"
EVENT_MIRROR_META_SHEET = "_shift_helper_meta"
EVENT_MIRROR_SCHEMA_VERSION = 2
_WRITE_LOCK = Lock()
_DELETED_STATUS = "deleted"

_HEADERS: tuple[str, ...] = (
    "Дата останова",
    "Время останова",
    "№ ВЭУ",
    "Описание события",
    "Причина",
    "Действия персонала",
    "Исполнитель",
    "Дата пуска",
    "Время пуска",
    "Простой",
    "Кто внёс запись",
    "Потери",
)
_COLUMN_WIDTHS: tuple[float, ...] = (
    14,
    11,
    15,
    44,
    34,
    38,
    24,
    14,
    11,
    14,
    24,
    14,
)


@dataclass(frozen=True, slots=True)
class EventMirrorResult:
    path: Path
    pending_path: Path
    generated_at: datetime
    record_count: int


class EventMirrorWriteError(RuntimeError):
    """Raised when a valid candidate cannot replace the public mirror file."""

    def __init__(self, target: Path, pending: Path, cause: OSError) -> None:
        super().__init__(
            f"Не удалось обновить {target.name}; закройте файл в Excel и повторите операцию."
        )
        self.target = target
        self.pending = pending
        self.__cause__ = cause


def _format_actor(actor: object) -> str:
    if actor is None or str(actor) in {"system", "migration"}:
        return ""
    value = str(actor)
    if value == "local":
        return "Локальное рабочее место"
    if value.startswith("lan:"):
        parts = value.split(":")
        if len(parts) >= 3 and parts[1]:
            return parts[1]
    return value


def _authors(session: Session, events: list[Event]) -> dict[int, str]:
    if not events:
        return {}
    event_ids = [event.id for event in events]
    rows = session.execute(
        text(
            """
            SELECT event_id, actor, action, id
            FROM event_audit
            WHERE event_id IN :event_ids
              AND action IN ('create', 'baseline')
            ORDER BY event_id,
                     CASE action WHEN 'create' THEN 0 ELSE 1 END,
                     id
            """
        ).bindparams(event_ids=tuple(event_ids))
    ).mappings()
    result: dict[int, str] = {}
    for row in rows:
        event_id = int(row["event_id"])
        result.setdefault(event_id, _format_actor(row["actor"]))
    return result


def _event_row(event: Event, author: str) -> list[object]:
    downtime: timedelta | None = None
    if event.end_at is not None:
        downtime = max(event.end_at - event.start_at, timedelta(0))
    return [
        event.start_at.date(),
        event.start_at.time().replace(second=0, microsecond=0),
        event.asset_label,
        event.description,
        event.reason,
        event.actions,
        event.performer,
        event.end_at.date() if event.end_at else None,
        event.end_at.time().replace(second=0, microsecond=0) if event.end_at else None,
        downtime,
        author,
        None,
    ]


def _style_workbook(
    workbook: Workbook,
    events: list[Event],
    authors: dict[int, str],
    generated_at: datetime,
) -> None:
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Не удалось создать лист зеркала журнала событий.")
    sheet.title = EVENT_MIRROR_SHEET
    sheet.append(list(_HEADERS))

    for event in events:
        sheet.append(_event_row(event, authors.get(event.id, "")))

    header_fill = PatternFill("solid", fgColor="D9E5F6")
    header_font = Font(name="Arial", size=10, bold=True, color="172033")
    body_font = Font(name="Arial", size=10, color="172033")
    thin_side = Side(style="thin", color="CDD5DF")
    body_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = body_border
    sheet.row_dimensions[1].height = 34

    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 30
        for column_index in range(1, len(_HEADERS) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = body_font
            cell.border = body_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_index in (1, 2, 3, 8, 9, 10, 12):
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )

        sheet.cell(row=row_index, column=1).number_format = "dd.mm.yyyy"
        sheet.cell(row=row_index, column=2).number_format = "hh:mm"
        sheet.cell(row=row_index, column=8).number_format = "dd.mm.yyyy"
        sheet.cell(row=row_index, column=9).number_format = "hh:mm"
        sheet.cell(row=row_index, column=10).number_format = "[h]:mm"
        sheet.cell(row=row_index, column=12).number_format = "0.00"

    for column_index, width in enumerate(_COLUMN_WIDTHS, start=1):
        letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[letter].width = width

    last_row = max(1, sheet.max_row)
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = f"A1:L{last_row}"
    sheet.print_title_rows = "1:1"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_view.showGridLines = True

    meta = workbook.create_sheet(EVENT_MIRROR_META_SHEET)
    meta.sheet_state = "hidden"
    meta.append(["schemaVersion", EVENT_MIRROR_SCHEMA_VERSION])
    meta.append(["generatedAt", generated_at.isoformat(timespec="seconds")])
    meta.append(["source", "SQLite"])
    meta.append([])
    meta.append(["row", "id", "revision", "updatedAt"])
    for visual_index, event in enumerate(events, start=1):
        meta.append(
            [
                visual_index,
                event.id,
                event.revision,
                event.updated_at.isoformat(timespec="seconds"),
            ]
        )

    workbook.properties.title = "Журнал событий Shift-Helper"
    workbook.properties.subject = "Совместимая экспортная копия данных SQLite"
    workbook.properties.creator = "Shift-Helper"
    workbook.properties.modified = generated_at


def refresh_event_journal_mirror(engine: Engine, exports_directory: Path) -> EventMirrorResult:
    """Rebuild and atomically publish the approved-form journal mirror."""

    exports_directory.mkdir(parents=True, exist_ok=True)
    target = exports_directory / EVENT_MIRROR_FILENAME
    pending = exports_directory / f".{EVENT_MIRROR_FILENAME}.pending.xlsx"
    generated_at = datetime.now().replace(microsecond=0)

    with _WRITE_LOCK:
        statement = (
            select(Event)
            .where(Event.status != _DELETED_STATUS)
            .order_by(Event.start_at.asc(), Event.id.asc())
        )
        with Session(engine) as session:
            events = list(session.scalars(statement))
            authors = _authors(session, events)
            workbook = Workbook()
            _style_workbook(workbook, events, authors, generated_at)
            workbook.save(pending)

        verifier = load_workbook(pending, read_only=True, data_only=False)
        try:
            if EVENT_MIRROR_SHEET not in verifier.sheetnames:
                raise RuntimeError("Проверка зеркала не нашла основной лист ЖС.")
            if EVENT_MIRROR_META_SHEET not in verifier.sheetnames:
                raise RuntimeError("Проверка зеркала не нашла лист служебных метаданных.")
        finally:
            verifier.close()

        try:
            os.replace(pending, target)
        except OSError as exc:
            raise EventMirrorWriteError(target, pending, exc) from exc

    return EventMirrorResult(
        path=target,
        pending_path=pending,
        generated_at=generated_at,
        record_count=len(events),
    )

from datetime import date, datetime, time
from hashlib import sha256
from io import BytesIO, TextIOWrapper
from pathlib import Path

from openpyxl import Workbook, load_workbook

from shift_helper.cli import _configure_console_stream, main
from shift_helper.core.journal_reader import read_event_journal
from shift_helper.core.selection import event_filter_code, select_emergency_events


def _journal(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ЖС"
    ws.append(
        [
            "Номер",
            "Дата останова ВЭУ",
            "Время останова ВЭУ",
            "40",
            "Описание",
            "Причины возникновения",
            "Действия персонала",
            "Непосредственный исполнитель (ФИО)",
            "Дата пуска ВЭУ",
            "Время пуска ВЭУ2",
        ]
    )
    # Known accidental leading outlier.
    ws.append(
        [
            1,
            date(2026, 2, 20),
            time(9, 15),
            20,
            "Остановлена для проведения работ",
            "Работы",
            None,
            None,
            date(2026, 2, 20),
            time(9, 55),
        ]
    )
    ws.append(
        [
            2,
            date(2026, 1, 1),
            time(5, 6),
            31,
            "Авария",
            "Причина",
            None,
            None,
            date(2026, 1, 1),
            time(5, 7),
        ]
    )
    ws.append(
        [
            3,
            date(2026, 7, 29),
            time(6, 59),
            10,
            "Авария",
            "До окна",
            None,
            None,
            date(2026, 7, 29),
            time(7, 1),
        ]
    )
    ws.append(
        [
            4,
            date(2026, 7, 29),
            time(7, 0),
            57,
            "Авария - SafSysChainBroken",
            "Низкое напряжение",
            None,
            None,
            None,
            None,
        ]
    )
    ws.append(
        [
            5,
            date(2026, 7, 29),
            time(8, 0),
            54,
            "Остановлена для работ",
            "ТО",
            None,
            None,
            date(2026, 7, 29),
            time(17, 0),
        ]
    )
    ws.append(
        [
            6,
            date(2026, 7, 30),
            time(6, 59),
            58,
            "Ошибка в работе преобразователя",
            "Потеря связи",
            None,
            None,
            date(2026, 7, 30),
            time(7, 0),
        ]
    )
    ws.append(
        [
            7,
            date(2026, 7, 30),
            time(7, 0),
            62,
            "Авария",
            "После окна",
            None,
            None,
            date(2026, 7, 30),
            time(7, 5),
        ]
    )
    wb.save(path)


def _template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Основные данные"
    ws["B1"] = "Другой лист не изменять"
    ws = wb.create_sheet("Аварийные отключения ЛЭП")
    ws.merge_cells("B1:F1")
    ws["B1"] = "Аварийные отключения ЛЭП и оборудования на 30.07.2026 Кочубеевская ВЭС"
    headers = (
        "Диспетчерское наименование ЛЭП, ВЭУ и оборудования",
        "Дата, время отключения",
        "Причина",
        "Работа защит",
        "Дата, время включения в работу",
    )
    for column, value in enumerate(headers, start=2):
        ws.cell(3, column).value = value
        ws.cell(4, column).value = "sample"
    wb.save(path)


def _digest(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def test_filter_order_matches_vba_contract() -> None:
    assert event_filter_code("Авария", "") == "skip.empty_reason"
    assert event_filter_code("Остановлена для работ", "Причина") == "skip.maintenance_context"
    assert event_filter_code("Ошибка в работе ПЧ", "Причина") is None
    assert event_filter_code("ВЭУ в работе", "Причина") == "skip.in_operation_context"


def test_read_select_write_and_diagnostics(tmp_path: Path) -> None:
    journal = tmp_path / "journal.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    diagnostics = tmp_path / "diagnostics"
    _journal(journal)
    _template(template)
    source_before = _digest(journal)

    result = main(
        [
            "build-emergency-report",
            "--journal",
            str(journal),
            "--template",
            str(template),
            "--report-date",
            "2026-07-30",
            "--output",
            str(output),
            "--diagnostics",
            str(diagnostics),
        ]
    )
    assert result == 0
    assert _digest(journal) == source_before
    assert output.is_file()
    assert (diagnostics / "event-selection.csv").is_file()
    assert (diagnostics / "validation.json").is_file()

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["Аварийные отключения ЛЭП"]
        assert ws["B4"].value == "ВЭУ №57"
        assert ws["B5"].value == "ВЭУ №58"
        assert ws["B6"].value is None
        assert ws["C4"].value == datetime(2026, 7, 29, 7, 0)
        assert ws["F4"].value is None
        assert wb["Основные данные"]["B1"].value == "Другой лист не изменять"
    finally:
        wb.close()


def test_actual_header_warning_is_non_blocking(tmp_path: Path) -> None:
    journal = tmp_path / "journal.xlsx"
    _journal(journal)
    result = read_event_journal(journal)
    assert any(issue.code == "journal.header.asset_mismatch" for issue in result.issues)
    assert 2 in result.ignored_rows
    selection = select_emergency_events(result.events, date(2026, 7, 30))
    assert [event.asset_number for event in selection.selected_events] == [57, 58]


def test_packaged_console_stream_supports_russian_help() -> None:
    buffer = BytesIO()
    stream = TextIOWrapper(buffer, encoding="cp1252")
    _configure_console_stream(stream)
    stream.write("Журнал событий")
    stream.flush()
    assert buffer.getvalue().decode("utf-8") == "Журнал событий"

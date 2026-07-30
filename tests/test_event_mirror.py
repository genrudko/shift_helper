from datetime import date, time, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from shift_helper import create_app
from shift_helper.event_mirror import (
    EVENT_MIRROR_FILENAME,
    EVENT_MIRROR_META_SHEET,
    EVENT_MIRROR_SHEET,
    EventMirrorWriteError,
    refresh_event_journal_mirror,
)

EXPECTED_HEADERS = (
    "Дата останова",
    "Время останова",
    "№ ВЭУ",
    "Описание события",
    "Причина",
    "Действия персонала",
    "Исполнитель",
    "Дата пуска",
    "Время пуска",
    "Простой",
    "Кто внёс запись",
    "Потери",
)


def _event_form() -> dict[str, str]:
    return {
        "start_at": "2026-07-29T11:30",
        "asset_label": "17",
        "event_type": "rotor_limit",
        "description": "Проверка зеркала журнала",
        "reason": "Повышенная вибрация",
        "actions": "Информация передана сменному персоналу",
        "performer": "Иванов И.И.",
        "error_codes": "214",
        "rotor_limit": "0,80",
        "include_in_report": "on",
    }


def test_event_mirror_tracks_approved_form_patch_and_pusk(tmp_path: Path) -> None:
    app = create_app(testing=True, data_root=tmp_path)
    client = app.test_client()
    mirror_path = tmp_path / "exports" / EVENT_MIRROR_FILENAME

    assert mirror_path.is_file()
    initial_health = client.get("/health").get_json()
    assert initial_health["eventMirror"]["status"] == "ok"
    assert initial_health["eventMirror"]["recordCount"] == 0

    created = client.post("/events/new", data=_event_form())
    assert created.status_code == 302
    assert created.headers["X-Shift-Helper-Event-Mirror"] == "ok"

    workbook = load_workbook(mirror_path, data_only=False)
    try:
        assert workbook.sheetnames == [EVENT_MIRROR_SHEET, EVENT_MIRROR_META_SHEET]
        assert EVENT_MIRROR_SHEET == "ЖС"
        sheet = workbook[EVENT_MIRROR_SHEET]
        meta = workbook[EVENT_MIRROR_META_SHEET]
        assert meta.sheet_state == "hidden"
        assert sheet.freeze_panes == "D2"
        assert sheet.auto_filter.ref == "A1:L2"
        assert tuple(cell.value for cell in sheet[1]) == EXPECTED_HEADERS
        assert sheet["A2"].value.date() == date(2026, 7, 29)
        assert sheet["B2"].value == time(11, 30)
        assert sheet["C2"].value == "17"
        assert sheet["D2"].value == "Проверка зеркала журнала"
        assert sheet["E2"].value == "Повышенная вибрация"
        assert sheet["F2"].value == "Информация передана сменному персоналу"
        assert sheet["G2"].value == "Иванов И.И."
        assert sheet["H2"].value is None
        assert sheet["I2"].value is None
        assert sheet["J2"].value is None
        assert sheet["K2"].value == "Локальное рабочее место"
        assert sheet["L2"].value is None
        assert sheet["A2"].number_format == "dd.mm.yyyy"
        assert sheet["B2"].number_format == "hh:mm"
        assert sheet["H2"].number_format == "dd.mm.yyyy"
        assert sheet["I2"].number_format == "hh:mm"
        assert sheet["J2"].number_format == "[h]:mm"
        assert meta["B1"].value == 2
        assert meta["B6"].value == 1
        assert meta["C6"].value == 1
    finally:
        workbook.close()

    patched = client.patch(
        "/events/api/v3/records/1",
        json={
            "revision": 1,
            "changes": {"description": "Описание обновлено"},
        },
    )
    assert patched.status_code == 200
    assert patched.headers["X-Shift-Helper-Event-Mirror"] == "ok"

    workbook = load_workbook(mirror_path, data_only=False)
    try:
        sheet = workbook[EVENT_MIRROR_SHEET]
        meta = workbook[EVENT_MIRROR_META_SHEET]
        assert sheet["D2"].value == "Описание обновлено"
        assert meta["C6"].value == 2
    finally:
        workbook.close()

    closed = client.patch(
        "/events/api/v3/records/1",
        json={
            "revision": 2,
            "changes": {"endAt": "2026-07-29T13:15"},
        },
    )
    assert closed.status_code == 200
    assert closed.headers["X-Shift-Helper-Event-Mirror"] == "ok"

    workbook = load_workbook(mirror_path, data_only=False)
    try:
        sheet = workbook[EVENT_MIRROR_SHEET]
        meta = workbook[EVENT_MIRROR_META_SHEET]
        assert sheet.max_row == 2
        assert sheet["H2"].value.date() == date(2026, 7, 29)
        assert sheet["I2"].value == time(13, 15)
        assert sheet["J2"].value == timedelta(hours=1, minutes=45)
        assert meta["C6"].value == 3
    finally:
        workbook.close()

    final_health = client.get("/health").get_json()
    assert final_health["eventMirror"]["status"] == "ok"
    assert final_health["eventMirror"]["recordCount"] == 1
    assert final_health["eventMirror"]["lastError"] is None
    assert not (tmp_path / "exports" / f".{EVENT_MIRROR_FILENAME}.pending.xlsx").exists()


def test_event_mirror_preserves_pending_candidate_when_target_is_locked(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    app = create_app(testing=True, data_root=tmp_path)
    engine = app.extensions["shift_helper_database_engine"]
    exports = tmp_path / "exports"

    def fail_replace(_source: Path, _target: Path) -> None:
        raise PermissionError("target is open")

    monkeypatch.setattr("shift_helper.event_mirror.os.replace", fail_replace)

    with pytest.raises(EventMirrorWriteError) as error:
        refresh_event_journal_mirror(engine, exports)

    assert "закройте файл в Excel" in str(error.value)
    assert error.value.pending.is_file()
    assert error.value.target == exports / EVENT_MIRROR_FILENAME

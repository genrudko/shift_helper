from __future__ import annotations

import hashlib
import zipfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from shift_helper.core.selection import event_filter_code
from shift_helper.core.selection import report_window as legacy_report_window
from shift_helper.core.workbook_contract import (
    INPUT_OUTAGES,
    INPUT_SHEETS,
    PREP_SHEET,
    REPORT_DATE_CELL,
    REPORT_OFFSET_CELL,
    REPORT_SHEETS,
    WTG_COUNT,
    WTG_STATUSES,
    available_power_mw,
    average_load_mw,
    plan_to_elapsed_days_kwh,
    remaining_month_hours,
    report_window,
    required_remaining_mean_power_kw,
    shifted_report_timestamp,
)
from shift_helper.extension_builder_payload import _TEMPLATE_ENTRY_SHA256, _template_bytes

ROOT = Path(__file__).resolve().parents[1]


def test_excel_client_owns_exact_shared_workbook_contract() -> None:
    assert PREP_SHEET == "Подготовка рапорта"
    assert REPORT_DATE_CELL == "B3"
    assert REPORT_OFFSET_CELL == "B6"
    assert INPUT_OUTAGES == "Ввод - Аварийные отключения"
    assert REPORT_SHEETS == (
        "Основные данные",
        "Аварийные отключения ЛЭП",
        "Команды по внешней инициативе",
        "Нарушения ОТиПБ + Экология",
        "Состояние ВЭУ",
        "Запланированные работы",
        "Дефекты оборудования",
    )
    assert len(INPUT_SHEETS) == len(REPORT_SHEETS) == 7
    assert WTG_COUNT == 84
    assert WTG_STATUSES == ("Работа", "Останов", "Авария", "Ремонт")
    assert sum((18, 58, 2, 6)) == WTG_COUNT


def test_accepted_main_calculations() -> None:
    assert average_load_mw(1_198_238) == pytest.approx(49.9265833333)
    assert required_remaining_mean_power_kw(
        27_830_685,
        14_478_050,
        date(2026, 7, 30),
    ) == pytest.approx(278_179.8958333333)
    assert available_power_mw(2.5, 0.75) == pytest.approx(1.75)
    assert available_power_mw(1.0, 2.5) == 0.0
    assert remaining_month_hours(date(2026, 7, 30)) == 48
    assert required_remaining_mean_power_kw(100, 101, date(2026, 7, 30)) == -1.0
    assert plan_to_elapsed_days_kwh(31_000, date(2026, 7, 30)) == pytest.approx(29_000)


def test_report_window_and_offset_match_calc_contract() -> None:
    report_date = date(2026, 7, 30)
    window = report_window(report_date)
    legacy_start, legacy_end = legacy_report_window(report_date)
    assert window.start == legacy_start == datetime(2026, 7, 29, 7, 0)
    assert window.end == legacy_end == datetime(2026, 7, 30, 7, 0)
    assert window.contains(datetime(2026, 7, 29, 7, 0))
    assert not window.contains(datetime(2026, 7, 30, 7, 0))

    source = datetime(2026, 7, 30, 8, 15)
    shifted = shifted_report_timestamp(source, -3)
    assert source == datetime(2026, 7, 30, 8, 15)
    assert shifted == datetime(2026, 7, 30, 5, 15)


def test_emergency_filter_is_not_replaced_by_keyword_guessing() -> None:
    assert event_filter_code("Ошибка в работе преобразователя", "Сработка защиты") is None
    assert event_filter_code("В работе", "Контроль") == "skip.in_operation_context"
    assert event_filter_code("Остановлена для работ", "ТО") == "skip.maintenance_context"
    assert event_filter_code("Аварийное отключение", "-") == "skip.empty_reason"


def test_approved_report_template_rebuild_preserves_every_exact_member() -> None:
    payload = _template_bytes(ROOT)
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        assert set(archive.namelist()) == set(_TEMPLATE_ENTRY_SHA256)
        for name, expected in _TEMPLATE_ENTRY_SHA256.items():
            assert hashlib.sha256(archive.read(name)).hexdigest() == expected
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    try:
        assert tuple(workbook.sheetnames) == REPORT_SHEETS
    finally:
        workbook.close()


def test_excel_vba_source_is_addin_owned_and_has_no_active_x_or_powershell() -> None:
    source_dir = ROOT / "packaging" / "excel_addin" / "vba"
    all_source = "\n".join(
        path.read_text(encoding="ascii") for path in sorted(source_dir.glob("*.bas"))
    )
    lowered = all_source.casefold()
    assert "outlook.application" in lowered
    assert "powershell" not in lowered
    assert "microsoft date and time picker" not in lowered
    assert "activex" not in lowered
    assert "sh_extractembeddedreporttemplate" in lowered
    assert "public function sh_reportdatecell" not in lowered

    report = (source_dir / "modShiftHelperReport.bas").read_text(encoding="ascii")
    assert "GetOpenFilename" not in report
    assert "GetSaveAsFilename" in report
    assert "shift_helper_report_template.xlsx" not in report
